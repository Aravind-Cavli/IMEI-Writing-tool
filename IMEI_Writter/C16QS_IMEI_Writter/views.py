import django
django.setup()
from django.shortcuts import render,HttpResponse
from .models import variants,hardware,Socket_Mapping
from django.urls import reverse
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.http import StreamingHttpResponse
import socketio
from django.contrib import messages



# Create your views here.

import os
import re
import sys
import time
import serial
import openpyxl
import multiprocessing
from datetime import date
from openpyxl.styles import PatternFill
from concurrent.futures import ProcessPoolExecutor
import json
from channels.generic.websocket import AsyncWebsocketConsumer

class MyConsumer(AsyncWebsocketConsumer):
    async def connect(self):
        await self.accept()

    async def disconnect(self, close_code):
        pass

    async def receive(self, text_data):
        data = json.loads(text_data)
        message = data['message']
         # process the message as needed

        # send a response back
        await self.send(text_data=json.dumps({'message': 'Response message'}))



def iccid_Check():
    ret = -1
    try:
        ser.write(("AT+ICCID\r\n").encode("utf-8"))
        response = ser.readlines()
        if b"OK\r\n" in response:
            for item in response:
                item = item.decode("utf-8").strip("\r\n")
                if "+ICCID:" in item:
                    iccid = list(re.findall("\-?\d+", item))
                    if "00000000000000000000" not in iccid[0]:
                        ret = iccid[0]
                        print("ICCID is : ", ret + ">" + "pass")
                    else:
                        ret = -1
        else:
            ret = -1
    except Exception as e:
        print(str(e) + ">fail \n")
        ret = -1

    return ret


def process_port_value(values):
    
    print(values)
    imei = values.split(",")[0]
    sn = values.split(",")[1]
    com_port = values.split(",")[2]
    variant_selected = values.split(",")[3]
    hw_version_selected = values.split(",")[4]
    username = values.split(",")[5]
    name= values.split(",")[6]
    
    lock = multiprocessing.Lock()

    startTest(com_port, imei, sn, variant_selected, hw_version_selected, username, lock)
   
    return imei, sn, com_port,name


def startTest(
    com_port, imei, sn, variant_selected, hw_version_selected, username,lock,
):
    #global message
    
    global IMEISN_STATUS, ESIM_STATUS, HW_VERSION_WRITE_STATUS, PN_WRITE_STATUS, PERSON_INCHARGE
    IMEISN_STATUS = "NA"
    ESIM_STATUS = "NA"
    HW_VERSION_WRITE_STATUS = "NA"
    PN_WRITE_STATUS = "NA"
    return_msg=""
    global ser, pass_fail_flag
    pass_fail_flag = True
    
    try:
        if IMEI_Write(
            com_port, imei, sn, variant_selected, hw_version_selected, username,
        ):
            if pass_fail_flag:
                try:
                    lock.acquire()
                    desktop = os.path.expanduser("~\\Desktop")
                    file_MasterExcel_path = (
                        desktop
                        + "\\DO_NOT DELETE\\MasterList\\C16QS_MasterList"
                        + ".xlsx"
                    )
                    book = openpyxl.load_workbook(file_MasterExcel_path)

                    sheet = book[variant_selected]

                    max_row = sheet.max_row
                    row_count = 1
                    today = date.today().strftime("%d-%m-%Y")
                    timestamp = time.strftime("%H-%M-%S")
                    PERSON_INCHARGE = username

                    sheet.cell(row=max_row + row_count, column=1).value = today
                    sheet.cell(row=max_row + row_count, column=2).value = timestamp
                    sheet.cell(
                        row=max_row + row_count, column=3
                    ).value = variant_selected
                    sheet.cell(row=max_row + row_count, column=4).value = (
                        imei + "/" + sn
                    )
                    sheet.cell(
                        row=max_row + row_count, column=5
                    ).value = Hardware_version
                    sheet.cell(
                        row=max_row + row_count, column=6
                    ).value = firmware_release
                    sheet.cell(row=max_row + row_count, column=7).value = IMEISN_STATUS
                    sheet.cell(
                        row=max_row + row_count, column=8
                    ).value = HW_VERSION_WRITE_STATUS
                    sheet.cell(
                        row=max_row + row_count, column=9
                    ).value = PN_WRITE_STATUS
                    sheet.cell(row=max_row + row_count, column=10).value = ESIM_STATUS
                    sheet.cell(
                        row=max_row + row_count, column=11
                    ).value = PERSON_INCHARGE

                    book.save(file_MasterExcel_path)
                    book.close()
                    lock.release()
                   
                    print(
                        ".....................COMPLETED...........................>pass\n"
                    )
                except:
                    
                    print("Excel Write   :FAIL>fail\n")
                    lock.release()
    except serial.SerialException as e:
        print(str(e) + ">fail \n")
     

def IMEI_Write(com_port, IMEI, SN, variant_selected, hw_version_selected, username,):
    global firmware_release, Hardware_version, part_number, Build_Date, IMEISN_STATUS, HW_VERSION_WRITE_STATUS, PN_WRITE_STATUS, ser, ESIM_STATUS,message
    flag = False
    try:
        ser = serial.Serial(com_port, 115200, timeout=0.25)
        ser.write(("AT\r\n").encode("utf-8"))
        response = ser.readlines()

        if b"OK\r\n" in response:
            
            print("Boot          :PASS"+IMEI+">pass\n")
           #msg_display("Boot          :PASS>pass\n")
           
            
            ser.write(("AT^HWVER=" + hw_version_selected + "\r\n").encode("utf-8"))
            response = ser.readlines()

            if b"OK\r\n" in response:
                
                #message.append("HW Ver write  :PASS>pass\n")
                #print(message)
                print("HW Ver write  :PASS>"+IMEI+"/"+SN+"pass\n")
                msg=JsonResponse({'message': 'View function returned value:' })
                HW_VERSION_WRITE_STATUS = "PASS"
            else:
                
                print("C16QS Hardware Version Write Error>fail\n")
                
                message.append("C16QS Hardware Version Write Error>fail\n")
                print(message)
                

                flag = False
                ser.close()
                return flag,msg

            ser.write(("AT^HWPN=" + variant_selected + "\r\n").encode("utf-8"))
            response = ser.readlines()

            if b"OK\r\n" in response:
                msg="HW PN WRITE : PASS"
                print("HW PN Write   :PASS>pass\n")
                #message.append("HW PN Write   :PASS>pass\n")
                #print(message)
                PN_WRITE_STATUS = "PASS"
            else:
                print("C16QS Hardware Part Number Write Error>fail\n")
               
                flag = False
                ser.close()
                return flag

            ser.write(("ATI\r\n").encode("utf-8"))
            response = ser.readlines()

            for item in response:
                item_d = item.decode("utf-8").strip("\r\n")
                if "Firmware Release:" in item_d:
                    firmware_release_list = item_d.split(":")
                    firmware_release = firmware_release_list[-1]
                elif "HW Version:" in item_d:
                    Hardware_version_list = item_d.split(":")
                    Hardware_version = Hardware_version_list[-1]
                elif "Part Number:" in item_d:
                    part_number_list = item_d.split(":")
                    part_number = part_number_list[-1]
                elif "Build Date:" in item_d:
                    Build_Date_list = item_d.split(":")
                    Build_Date = Build_Date_list[-1]

            if (
                variant_selected == "C16QS-EA-GNAH"
                or variant_selected == "C16QS-NA-GNAH"
                or variant_selected == "C16QS-EA-S00H"
                or variant_selected == "C16QS-NA-S00H"
                or variant_selected == "C16QS-AN-S00H"
                or variant_selected == "C16QS-AN-GNAH"
                or variant_selected == "C16QS-LA-S00H"
                or variant_selected == "C16QS-LA-GNAH"
                or variant_selected == "C16QS-WW-S00H"
                or variant_selected == "C16QS-WW-GNAH"
                or variant_selected == "C16QS-WW-GNAN"
            ):
                ret = iccid_Check()
                if ret != -1:
                    #list(message).append("ESIM Check    :PASS" + " , " + ret + ">" + "pass" + "\n")
                    print("ESIM Check    :PASS" + " , " + ret + ">" + "pass" + "\n")
                    ESIM_STATUS = "PASS"
                else:
                    print("ESIM Check    :FAIL>fail\n")
                    flag = False
                    ser.close()
                    return flag

            ser.write(('AT$QCCGSN="imei",' + IMEI + "\r\n").encode("utf-8"))
            response = ser.readlines()
            
            ser.write(("AT+CGSN=1\r\n").encode("utf-8"))
            response = ser.readlines()

            if '+CGSN: "' + IMEI + '"' != response[1].decode("utf-8").strip("\r\n"):
                print("IMEI Error>fail\n")
               # message.append("IMEI Error>fail\n")
                flag = False
                ser.close()
                return flag

            ser.write(('AT$QCCGSN="sn",' + SN + "\r\n").encode("utf-8"))
            response = ser.readlines()

            ser.write(("AT+CGSN\r\n").encode("utf-8"))
            response = ser.readlines()

            
            if SN != response[1].decode("utf-8").strip("\r\n"):
                print("SN Error>fail\n")
                flag = False
                ser.close()
                return flag 
            msg='imei sn write pass'
            print("IMEI_SN write :PASS>pass\n")
            IMEISN_STATUS = "PASS"
            ser.close()
            flag = True
        else:
            print(
                "Serial is Not Working, pls check if module is powered on or not.>fail\n"
            )
            ser.close()
            flag = False
    except serial.SerialException as e:
        # print(str(e) + ">fail \n")
        print("Failed to open COM port:",com_port + ">fail")

        flag = False
    return flag


   
def excel_creation():  
    global variant_list
    desktop = os.path.expanduser("~\\Desktop")
    file_MasterExcel_path = (
        desktop + "\\DO_NOT DELETE\\MasterList\\C16QS_MasterList" + ".xlsx"
    )

    if not os.path.exists(file_MasterExcel_path):
        Master_List = openpyxl.Workbook()
        Master_List["Sheet"].title = "C16QS-EA-GNAH"
        Master_List.create_sheet("C16QS-EA-GNAN")
        Master_List.create_sheet("C16QS-NA-GNAH")
        Master_List.create_sheet("C16QS-NA-GNAN")
        Master_List.create_sheet("C16QS-EA-S00N")
        Master_List.create_sheet("C16QS-EA-S00H")
        Master_List.create_sheet("C16QS-NA-S00N")
        Master_List.create_sheet("C16QS-NA-S00H")
        Master_List.create_sheet("C16QS-AN-S00N")
        Master_List.create_sheet("C16QS-AN-S00H")
        Master_List.create_sheet("C16QS-AN-GNAN")
        Master_List.create_sheet("C16QS-AN-GNAH")
        Master_List.create_sheet("C16QS-LA-S00N")
        Master_List.create_sheet("C16QS-LA-S00H")
        Master_List.create_sheet("C16QS-LA-GNAN")
        Master_List.create_sheet("C16QS-LA-GNAH")
        Master_List.create_sheet("C16QS-WW-S00N")
        Master_List.create_sheet("C16QS-WW-S00H")
        Master_List.create_sheet("C16QS-WW-GNAN")
        Master_List.create_sheet("C16QS-WW-GNAH")
        Master_List.create_sheet("C16QS-WW-GNBN")
        Master_List.create_sheet("C16QS-WW-GNBH")
        
        

        #variant=["C16QS-EA-S00N","C16QS-EA-S00H","C16QS-EA-GNAN","C16QS-EA-GNAH","C16QS-NA-S00N","C16QS-NA-S00H","C16QS-NA-GNAN","C16QS-NA-GNAH","C16QS-AN-S00N","C16QS-AN-S00H","C16QS-AN-GNAN","C16QS-AN-GNAH","C16QS-LA-S00N","C16QS-LA-S00H","C16QS-LA-GNAN","C16QS-LA-GNAH","C16QS-WW-S00N","C16QS-WW-S00H","C16QS-WW-GNAN","C16QS-WW-GNAH","C16QS-WW-GNBN","C16QS-WW-GNBH"]
        variant=variant_list
        
        for i in variant:
            Master_List[i].cell(row=1, column=1).value = "DATE"
            Master_List[i].cell(row=1, column=2).value = "TIME"
            Master_List[i].cell(row=1, column=3).value = "VARIANT"
            Master_List[i].cell(row=1, column=4).value = "IMEI_SN"
            Master_List[i].cell(row=1, column=5).value = "HW_VERSION"
            Master_List[i].cell(row=1, column=6).value = "FIRMWARE RELEASE"
            Master_List[i].cell(row=1, column=7).value = "IMEI_SN_WRITE"
            Master_List[i].cell(row=1, column=8).value = "HW_VER_WRITE"
            Master_List[i].cell(row=1, column=9).value = "PN_Write"
            Master_List[i].cell(row=1, column=10).value = "ESIM_CHECK"
            Master_List[i].cell(row=1, column=11).value = "PERSON_INCHARGE"

            for col in range(1, 12):
                cell_header = Master_List[i].cell(1, col)
                cell_header.fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                )


        Master_List.save(file_MasterExcel_path)
        Master_List.close()
    return True

    # imei = sys.argv[1]
    # sn = sys.argv[2]
    # port = sys.argv[3]
    # variant_selected = sys.argv[4]
    # hw_version_selected = sys.argv[5]
    # username = sys.argv[6]

    # # Create a multiprocessing Pool
    # pool = multiprocessing.Pool()

    # # Apply the processing function to each port value asynchronously
    # result = pool.map_async(
    #     process_port_value,
    #     [
    #         imei
    #         + ","
    #         + sn
    #         + ","
    #         + port
    #         + ","
    #         + variant_selected
    #         + ","
    #         + hw_version_selected
    #         + ","
    #         + username
    #     ],
    # )

    # # # Close the pool to prevent further tasks
    # # pool.close()

    # # Wait for the results and retrieve them
    # result.wait()

    # # Terminate the pool
    # pool.terminate()



def printing(name,ver):
        print(name,ver)


def variant_rendering(request):
    global variant_list,variant,username,hw_ver
    
    variant_menu=variants.objects.values()
    hw_version_list=hardware.objects.values_list()
    hw_version=hw_version_list[0][1]
    print(variant_menu)
    
    variant_list=[]
    for items in variant_menu:
        if 'variant' in items:
            variant_list.append(items['variant']) 
        
        
    print("\n",variant_list)
   
    #print(hw_version)
    if request.POST:      #Getting hw_version/username/variant from front-end
        hw_ver=request.POST.get("hardwareVersion")
        username=request.POST.get("userName")
        variant=request.POST.get("variant")
        print(hw_ver)
        print(username)
        print(variant)
        sockets=Socket_Mapping.objects.values()
        
        
        return render(request,"index.html",{'SOCKETS':sockets,'VARIANT':variant})
    return render(request ,"variant.html",{'variant_menu': variant_menu,'hw':hw_version})
    
    
def interface(request):
    return render(request,"index.html")


def socket_mapping(request):
    sockets=Socket_Mapping.objects.values()
    print(sockets)
    return render(request,"socket_mapping.html",{'sockets':sockets})


def my_python_function(request):
    print("excecuted")
    # Your Python function logic goes here
    result = "Python function executed successfully"
  
    return JsonResponse({'result': result})    

    # for items in sockets:
    #     sl=items[1]
    #     socket_id=items[2]
    #     comports=items[3]
        
    #     socket_data= {'socket':[{'SL':sl},{'SOCKET_ID':socket_id},{'COM':comports,}]}

        # print('\n',sl)
        # print(socket_id)
        # print(comports)
 
        #print(socket_data)
        #return render(request,"socket_mapping.html",socket_data)
    
    
def admin_button(request):
    admin_url = reverse('admin:index')  # This will generate the URL to the admin panel
    return render(request, 'admin_button.html', {'admin_url': admin_url})


'''def my_view(request):
    
    if request.method == 'POST':
        # Get data from the request
        data = json.loads(request.body)
        
        print(data)
        # Call your Python function with the data
        result = my_python_function(data['data_to_send'])

        # Return a JSON response with the result
        return JsonResponse({'result': result})'''
    




def see_hi():
    return "Aravind"

def process_data(request):
    
    global username,variant,hw_ver
    
    
    if request.method == "POST":
        pool=multiprocessing.Pool()
        imei_sn = json.loads(request.body.decode('utf-8'))["input_data"]
        com=json.loads(request.body.decode('utf-8'))["com_port"]
        print(com)
        #print(imei_sn)
        # Your view logic here, you can use 'input_data' in your processing
        if excel_creation():
            print('excel creation success')
            #sse(request,"hlooooooooooooo")
            try:
                print(request)
                print(imei_sn)
                parse =imei_sn.split()
                temp=parse[-1].split('/')
                imei = temp[0]
                sn = temp[1]
                print(imei)
                print(sn)
                ret = see_hi()
                
                # Apply the processing function to each port value asynchronously
                result = pool.map_async(
                    process_port_value,
                    [
                        imei
                        + ","
                        + sn
                        + ","
                        + com
                        + ","
                        + variant
                        + ","
                        + hw_ver
                        + ","
                        + username
                        + ","
                        + ret
                         
                        
                        
                    ],
                
                
                )
                
                 
                
                
                #result=pool.map_async(startTest,[com,imei,sn,var,hw_ver,username])
                #startTest(com,imei,sn,var,hw_ver,username)
                result.wait()
                #print(message)
                output=result.get()
                print(output)
                pool.terminate()
            except Exception as e:
                
                print(e)
            
            #
            # return JsonResponse({'message': 'View function returned value: '+str(output) })
        
        return JsonResponse({'message': 'View function called with data: ' + imei_sn})
    
    return JsonResponse({'message': 'View function called without data'})

def event_stream(message):
        
        
        count=0
        send_flag=True
        while send_flag:
            count+=1
            time.sleep(.2)  # Simulate delay
            
            yield f"data: {message}\n\n"  # Send real-time data
            print(message)
            if count==10:
                send_flag=False
                
            else:
                time.sleep(.2)



def sse(request):
    
    # def event_stream(HY):
        
    #     count=0
    #     send_flag=True
    #     while send_flag:
    #         count+=1
    #         time.sleep(.2)  # Simulate delay
            
    #         yield f"data: {'data'}\n\n"  # Send your real-time data
    #         if count==10:
    #             send_flag=False
                
    #         else:
    #             time.sleep(.2)
        # return event_stream()
    
    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    return response

